#!/usr/bin/env python3
"""
Generate funnel-dashboard/data.json from the "Claude Transition - MEA Ops" workbook.

The Funnel Dashboard (index.html) is a static page: it cannot call the live Google
Sheet, so it reads ./data.json instead. This script produces that file.

SECURITY: the source sheet has a Password column — it is intentionally NEVER written
to data.json. Do not publish the raw sheet to the web.

Usage:
    python3 generate-data.py path/to/Claude_Transition_MEA_Ops.xlsx
    # writes data.json next to this script   (tab: "EA Pilot 62")

Columns are resolved BY HEADER NAME (row 2), not fixed positions — the sheet gets
re-ordered periodically (e.g. Insight Call was split into Call 1 / Call 2), which
silently corrupted the old index-based mapping. Duplicate headers are disambiguated
by occurrence: there are two "Status" columns (account status = first, call status =
last) and two each of "Facilitator" / "Recording Link" (Call 1 = first). A hard
validation aborts if required columns are missing, so a reshuffle fails loudly
instead of writing zeros.
"""
import sys, os, re, json, datetime
try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

XLSX = sys.argv[1] if len(sys.argv) > 1 else "Claude_Transition_MEA_Ops.xlsx"
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.json")
TAB = "EA Pilot 62"

# Internal accounts (Alver / Jason / Norj) that sit in the "Confirmed Access" seat
# roster but are NOT pilot participants — always excluded from any Confirmed Access
# count or stat so seat totals reflect EAs only.
CONFIRMED_ACCESS_EXCLUDE = {
    "jason@getmagicnow.com",
    "norjielyn@gmail.com",
    "alverremolar@gmail.com",
}

wb = openpyxl.load_workbook(XLSX, data_only=True)
if TAB not in wb.sheetnames:
    sys.exit(f'Tab "{TAB}" not found. Available: {wb.sheetnames}')
ws = wb[TAB]
wsh = openpyxl.load_workbook(XLSX)[TAB]   # second load keeps hyperlink targets
rows = list(ws.iter_rows(values_only=True))
if len(rows) < 3:
    sys.exit("Sheet has no data rows.")

# --- resolve column indices by header name (row 2 = rows[1]) ---
hm = {}  # header text -> list of 0-based column indices (in sheet order)
for i, name in enumerate(rows[1]):
    if name not in (None, ""):
        hm.setdefault(str(name).strip(), []).append(i)

def one(name, occ=0):
    idxs = hm.get(name, [])
    return idxs[occ] if idxs and -len(idxs) <= occ < len(idxs) else None
def one_of(*names):
    # first matching header, tolerant of index 0 (can't use `a or b` — index 0 is falsy)
    for n in names:
        i = one(n)
        if i is not None:
            return i
    return None
def contains(sub):
    for name, idxs in hm.items():
        if sub.lower() in name.lower():
            return idxs[0]
    return None

status_cols = hm.get("Status", [])
COL = dict(
    name=one_of("Assistant Name", "Name"), hsEmail=one("HS Email"), dealCard=one("Deal Card Link"),
    client=one("Client Name"), al=one("Account Lead"),
    status=(status_cols[0] if status_cols else None),      # account status = first "Status"
    email=one("Magic Assistant Email"), revoke=one("Revoke Access Reason"),
    invited=one("Invited to Claude"), accessed=one("Accessed Asst Email"),
    confirmed=one("Claude Access confirmed?"), desktop=one("Claude Desktop app installed?"),
    cowork=one("Has Cowork usage?"), vertical=one("Verticals"),
    sessionNote=one("Cowork Session Notes"), reachType=one("Reachout Type"),
    emailReach=one("Email Reachout"), discordReach=one("Discord Reachout"),
    discordUN=one("Discord UN"), reachNotes=one("Reachout Notes"),
    callStatus=(status_cols[-1] if len(status_cols) >= 2 else None),  # call status = last "Status"
    callDate=(contains("Insight Call 1 Date") or contains("Insight Call Date")),
    call2Date=contains("Insight Call 2 Date"),
    facil=one("Facilitator", 0),                           # Call 1 facilitator = first
    call2Facil=one("Facilitator", 1),                      # Call 2 facilitator = second
    profile=one("EA Profile Link"),
    rec=one("Recording Link", 0),                          # Call 1 recording = first
    call2Rec=one("Recording Link", 1),                     # Call 2 recording = second
    insight=one("Insight Call Notes"),
    bonus=one("Bonus Filed?"), dateFiled=contains("Date Filed"),
    rate=one("Assistant Rate"), payout=contains("Payout"),
    gen=(contains("General Notes")),
)
required = ["name", "al", "status", "email", "invited", "cowork", "callStatus", "bonus"]
missing = [k for k in required if COL[k] is None]
if missing:
    sys.exit(f"Column mapping failed — headers not found for: {missing}. "
             f"The sheet layout changed; update generate-data.py. Headers seen: {sorted(hm)}")

def cell(r, key):
    i = COL[key]
    return r[i] if i is not None and i < len(r) else None
def v(x): return "" if x is None else str(x).strip()
def tb(x): return v(x).lower() == "true"
def link(sheet_row, key):
    i = COL[key]
    if i is None: return ""
    c = wsh.cell(row=sheet_row, column=i + 1)
    if c.hyperlink and c.hyperlink.target:
        return c.hyperlink.target.strip()
    val = v(c.value)
    return val if val.startswith("http") else ""
def clean_al(s):
    a = re.sub(r"\s+Bot\s+", " ", s or "")
    a = re.sub(r"\s*Kenreich\s*", " ", a)
    a = re.sub(r"\s+", " ", a).strip()
    return a if a and a != "N/A" else "Unassigned"
def call_cat(s):
    s = s or ""
    if re.search("Call 2", s, re.I): return "call2"
    if re.search("Call 1", s, re.I): return "call1"
    if re.search("Done|Completed", s, re.I): return "done"
    return ""

out = []
for ri, r in enumerate(rows):
    if ri < 2:                       # rows 1-2 are headers
        continue
    if cell(r, "name") in (None, ""):
        continue
    sheet_row = ri + 1               # openpyxl is 1-based
    email = v(cell(r, "email")); email = email if email and email != "N/A" else ""
    status = v(cell(r, "status")).upper()
    o = dict(
        name=v(cell(r, "name")), hsEmail=v(cell(r, "hsEmail")), dealCard=v(cell(r, "dealCard")),
        client=v(cell(r, "client")), al=clean_al(v(cell(r, "al"))), status=status, email=email,
        revokeReason=v(cell(r, "revoke")),
        invited=tb(cell(r, "invited")), accessed=tb(cell(r, "accessed")),
        confirmed=tb(cell(r, "confirmed")), desktop=tb(cell(r, "desktop")), cowork=tb(cell(r, "cowork")),
        vertical=v(cell(r, "vertical")), sessionNote=v(cell(r, "sessionNote")),
        reachType=v(cell(r, "reachType")), emailReach=tb(cell(r, "emailReach")),
        discordReach=tb(cell(r, "discordReach")), discordUN=v(cell(r, "discordUN")),
        reachNotes=v(cell(r, "reachNotes")),
        callDate=v(cell(r, "callDate")), callCat=call_cat(v(cell(r, "callStatus"))), facil=v(cell(r, "facil")),
        call2Date=v(cell(r, "call2Date")), call2Facil=v(cell(r, "call2Facil")),
        profile=link(sheet_row, "profile"), rec=link(sheet_row, "rec"), call2Rec=link(sheet_row, "call2Rec"),
        insight=v(cell(r, "insight")), bonus=tb(cell(r, "bonus")), dateFiled=v(cell(r, "dateFiled")),
        rate=v(cell(r, "rate")), payout=v(cell(r, "payout")), gen=v(cell(r, "gen")),
        # Password column intentionally omitted.
    )
    o["reached"] = o["emailReach"] or o["discordReach"]
    o["inPilot"] = bool(email) and status not in ("CHURNED", "")
    # "inactive" = access was revoked / the assistant churned. Identified from the
    # "Revoke Access Reason" column (or a CHURNED status). Usage insights for these
    # people may still be shown, but must be labelled inactive (see index.html drawer).
    o["inactive"] = bool(o["revokeReason"]) or status == "CHURNED"
    out.append(o)

# --- Cowork usage aggregates (from "Cowork Users"; the raw "Cowork Sessions" prompt
#     text is NEVER read/exported) ---
usage = {}
if "Cowork Users" in wb.sheetnames:
    urows = list(wb["Cowork Users"].iter_rows(values_only=True))
    uh = {str(n).strip(): i for i, n in enumerate(urows[0]) if n not in (None, "")}
    def unum(row, name):
        i = uh.get(name)
        if i is None or i >= len(row) or row[i] in (None, ""): return 0.0
        try: return float(row[i])
        except Exception: return 0.0
    ei = uh.get("user_email")
    # The tab can carry MULTIPLE rows per user — an authoritative pull WITH cost/tokens
    # plus a supplementary pull that reports sessions/days but 0 cost/0 tokens. Collect
    # every row per email, then keep the richest one (most cost, then tokens, then
    # sessions) so a stray 0-cost duplicate never zeroes out a top user's real spend.
    by_email = {}
    for row in urows[1:]:
        em = str(row[ei]).strip().lower() if ei is not None and ei < len(row) and row[ei] else ""
        if not em: continue
        # "total_cost_usd" was dropped from this tab in the Jul 30 refresh; unum() returns
        # 0.0 when a header is absent, so `cost` stays in the contract (index.html hides a
        # zero) while the newer engagement columns below carry the signal instead.
        # avg_session_duration_seconds is deliberately NOT exported: its values (275h, 195h)
        # are implausible as averages, so publishing them would be misleading.
        by_email.setdefault(em, []).append(
            {"sessions": int(unum(row, "num_sessions")),
             "active_days": int(unum(row, "active_days")),
             "cost": round(unum(row, "total_cost_usd"), 2),
             "tokens": int(unum(row, "total_tokens")),
             "skills": int(unum(row, "num_skill_activations")),
             "distinctSkills": int(unum(row, "distinct_skills_used")),
             "hooks": int(unum(row, "num_hook_executions")),
             "toolOk": int(unum(row, "num_tool_successes"))})
    for em, recs in by_email.items():
        usage[em] = max(recs, key=lambda r: (r["cost"], r["tokens"], r["sessions"]))

# flag usage rows that belong to inactive (churned/revoked) assistants, so the
# dashboard can surface their insights clearly labelled as inactive.
inactive_emails = {o["email"].lower() for o in out if o.get("inactive") and o.get("email")}
for em in usage:
    usage[em]["inactive"] = em in inactive_emails

# --- "Cowork Sessions": usage "as of" date + per-user session counts.
#     ONLY session_date / user_email / session_id are read. The "all_prompts" column
#     (raw prompt text) is NEVER touched — see the module docstring.
usage_as_of = ""
sess_ids = {}      # email -> set of distinct session_id
sess_last = {}     # email -> most recent session_date
if "Cowork Sessions" in wb.sheetnames:
    sws = wb["Cowork Sessions"]
    srows = sws.iter_rows(values_only=True)
    shdr = next(srows, ())
    sidx = {str(n).strip(): i for i, n in enumerate(shdr) if n not in (None, "")}
    di, emi, sii = sidx.get("session_date"), sidx.get("user_email"), sidx.get("session_id")
    ds = []
    for row in srows:
        d = str(row[di])[:10] if di is not None and di < len(row) and row[di] else ""
        em = (str(row[emi]).strip().lower()
              if emi is not None and emi < len(row) and row[emi] else "")
        if d: ds.append(d)
        if em:
            if sii is not None and sii < len(row) and row[sii]:
                sess_ids.setdefault(em, set()).add(str(row[sii]))
            if d: sess_last[em] = max(sess_last.get(em, ""), d)
    if ds: usage_as_of = max(ds)

# --- "Confirmed Access": the authoritative Claude Team seat roster (Name/Email/Role/
#     Status/Seat Tier). Used to stamp each assistant with whether they actually hold a
#     seat, so a stale "Claude Access confirmed?" checkbox can be caught.
seats = {}
if "Confirmed Access" in wb.sheetnames:
    caws = wb["Confirmed Access"]
    carows = list(caws.iter_rows(values_only=True))
    cah = {str(n).strip(): i for i, n in enumerate(carows[0]) if n not in (None, "")}
    ce, cs, ct = cah.get("Email"), cah.get("Status"), cah.get("Seat Tier")
    for row in carows[1:]:
        em = str(row[ce]).strip().lower() if ce is not None and ce < len(row) and row[ce] else ""
        if not em: continue
        if em in CONFIRMED_ACCESS_EXCLUDE: continue   # internal accounts, never counted
        seats[em] = {"status": v(row[cs]) if cs is not None and cs < len(row) else "",
                     "tier": v(row[ct]) if ct is not None and ct < len(row) else ""}

roster_emails = {o["email"].lower() for o in out if o.get("email")}
for o in out:
    em = o["email"].lower()
    s = seats.get(em)
    # NOTE: the tab is keyed on Email, not Name — two rows carry an email with a blank
    # Name, so matching on Name would silently drop real seats.
    o["seat"] = bool(s) and s["status"].lower() == "active"
    o["seatStatus"] = s["status"] if s else ""
    o["seatTier"] = s["tier"] if s else ""
    o["sessionCount"] = len(sess_ids.get(em, ()))
    o["lastSession"] = sess_last.get(em, "")

# --- data-quality flags surfaced on the dashboard's "Attention flags" card ---
data_flags = []
no_agg = [o["name"] for o in out if o["sessionCount"] > 0 and o["email"].lower() not in usage]
if no_agg:
    data_flags.append(
        f"Session activity but no row in the Cowork Users aggregate: {', '.join(no_agg)}. "
        f"Their sessions are visible in Cowork Sessions, so the usage totals understate them.")
seat_gap = [o["name"] for o in out if o["confirmed"] and not o["seat"]
            and o["status"] not in ("CHURNED", "PAUSED")]
if seat_gap:
    data_flags.append(
        f"Marked access-confirmed but holds no Claude Team seat in Confirmed Access: "
        f"{', '.join(seat_gap)}.")
cow_gap = [o["name"] for o in out if o["cowork"] and o["sessionCount"] == 0]
if cow_gap:
    data_flags.append(f"Flagged 'Has Cowork usage?' but 0 sessions on record: {', '.join(cow_gap)}.")
cow_miss = [o["name"] for o in out if not o["cowork"] and o["sessionCount"] > 0]
if cow_miss:
    data_flags.append(f"Has sessions on record but 'Has Cowork usage?' is unchecked: {', '.join(cow_miss)}.")
pend = [o["name"] for o in out if o.get("seatStatus", "").lower() == "pending"]
if pend:
    data_flags.append(f"Claude Team seat invite still pending acceptance: {', '.join(pend)}.")
orphan = sorted(set(usage) - roster_emails)
if orphan:
    data_flags.append(f"{len(orphan)} usage row(s) with no matching assistant in the roster.")

# --- semantic sanity check: refuse to overwrite good data with corrupt data ---
def metrics(objs):
    return dict(n=len(objs),
                invited=sum(1 for o in objs if o.get("invited")),
                pipeline=sum(1 for o in objs if o.get("callCat")),
                bonus=sum(1 for o in objs if o.get("bonus")),
                reached=sum(1 for o in objs if o.get("reached")))
cur = metrics(out)
if cur["n"] < 10:
    sys.exit(f"Sanity check failed: only {cur['n']} assistants parsed — not overwriting data.json.")
if cur["invited"] == 0:
    sys.exit("Sanity check failed: 0 invited — likely a column mis-map. Not writing data.json.")
try:
    with open(OUT, encoding="utf-8") as f:
        pj = json.load(f)
    prev = metrics(pj.get("assistants", pj) if isinstance(pj, dict) else pj)
except Exception:
    prev = None
if prev and prev["n"]:
    if cur["n"] < prev["n"] * 0.7:
        sys.exit(f"Sanity check failed: assistants dropped {prev['n']} -> {cur['n']} (>30%). Not writing.")
    for key in ("pipeline", "bonus", "reached"):
        if prev[key] > 0 and cur[key] == 0:
            sys.exit(f"Sanity check failed: '{key}' went {prev[key]} -> 0 (likely a sheet column shift). Not writing.")

payload = {
    "generatedAt": datetime.datetime.now().isoformat(timespec="seconds"),
    "usageAsOf": usage_as_of,
    "assistants": out,
    "usage": usage,
    # Claude Team seat reconciliation (from the "Confirmed Access" tab). "roster" counts
    # only seats held by someone on the EA Pilot 62 roster; the remainder are internal.
    "seats": {"total": len(seats),
              "active": sum(1 for s in seats.values() if s["status"].lower() == "active"),
              "pending": sum(1 for s in seats.values() if s["status"].lower() == "pending"),
              "roster": sum(1 for em in seats if em in roster_emails),
              "internal": sum(1 for em in seats if em not in roster_emails)},
    "dataFlags": data_flags,
}
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(payload, f, ensure_ascii=False, indent=1)

# --- append today's aggregate snapshot to history.json (durable, shared trend lines) ---
HIST = os.path.join(os.path.dirname(OUT), "history.json")
pilot = [o for o in out if o["inPilot"]]
def money(o):
    try: return float(re.sub(r"[^0-9.]", "", str(o["payout"])) or 0)
    except Exception: return 0.0
incpay = [o for o in out if o["bonus"] and not o["dateFiled"]]
k = dict(cow=sum(1 for o in pilot if o["cowork"]),
         notStarted=sum(1 for o in pilot if not o["cowork"]),
         callsToRun=sum(1 for o in out if o["callCat"] in ("call1", "call2")),
         incPayN=len(incpay), incAmt=round(sum(money(o) for o in incpay), 2),
         invited=cur["invited"], accessed=sum(1 for o in out if o["accessed"]),
         confirmed=sum(1 for o in out if o["confirmed"]), desktop=sum(1 for o in out if o["desktop"]),
         coworkTotal=sum(1 for o in out if o["cowork"]))
today = datetime.date.today().isoformat()
hist = []
try:
    with open(HIST, encoding="utf-8") as f: hist = json.load(f)
except Exception: pass
hist = [h for h in hist if h.get("date") != today] + [{"date": today, "k": k}]
hist = sorted(hist, key=lambda h: h.get("date", ""))[-120:]
with open(HIST, "w", encoding="utf-8") as f:
    json.dump(hist, f, ensure_ascii=False, indent=1)

print(f"Wrote {OUT} ({len(out)} assistants, {len(usage)} usage rows, passwords excluded) "
      f"+ history.json ({len(hist)} days) at {payload['generatedAt']}")
